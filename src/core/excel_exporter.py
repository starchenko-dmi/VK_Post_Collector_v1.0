# -*- coding: utf-8 -*-
"""Экспорт данных в Excel с обработкой длинных текстов"""
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
from datetime import datetime
from typing import List, Dict
from pathlib import Path


class ExcelExporter:
    def __init__(self, output_dir: str, logger):
        self.output_dir = output_dir
        self.logger = logger
        self.full_text_dir = Path(output_dir) / "полные_тексты"
        self.full_text_dir.mkdir(exist_ok=True)

    def export_posts(self, posts: List[Dict]):
        """Экспорт постов в Excel с обработкой длинных текстов и конвертацией дат"""
        # Генерируем имя файла с датой
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = Path(self.output_dir) / f"отчёт_{timestamp}.xlsx"

        # Создаём книгу Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Посты"

        # Заголовки
        headers = [
            "Группа_ID", "Группа_название", "Пост_ID", "Дата_публикации",
            "Текст_полный", "Лайки", "Репосты", "Комментарии", "Ссылка_на_пост"
        ]
        ws.append(headers)

        # Стили
        header_font = Font(bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        # Применяем стили к заголовкам
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.border = border

        # Данные
        for post in posts:
            # Обработка текста
            full_text = post['text']
            text_display = full_text[:32760] + "..." if len(full_text) > 32767 else full_text

            # Если текст длинный — сохраняем в отдельный файл
            if len(full_text) > 32767:
                safe_filename = f"пост_{post['group_id']}_{post['post_id'].replace('/', '_')}.txt"
                file_path = self.full_text_dir / safe_filename
                try:
                    with open(file_path, "w", encoding="utf-8") as f:
                        f.write(full_text)
                    text_display += f" [полный текст в файле: {safe_filename}]"
                    self.logger.warning(f"Текст поста {post['post_id']} сохранён в файл: {safe_filename}")
                except Exception as e:
                    self.logger.error(f"Ошибка сохранения полного текста: {e}")
                    text_display += " [ошибка сохранения полного текста]"

            # Конвертация даты для Excel (удаление временной зоны)
            post_date = post['date']
            if hasattr(post_date, 'tzinfo') and post_date.tzinfo is not None:
                # Конвертируем в локальное время без tzinfo (для корректного отображения в Excel)
                date_for_excel = post_date.replace(tzinfo=None)
            else:
                date_for_excel = post_date

            # Добавляем строку в Excel
            row = [
                post['group_id'],
                post['group_name'],
                post['post_id'],
                date_for_excel,  # ← ИСПРАВЛЕНО: дата без tzinfo
                text_display,
                post['likes'],
                post['reposts'],
                post['comments'],
                post['post_url']
            ]
            ws.append(row)

        # Автоматическая подстройка ширины колонок
        for col_idx, header in enumerate(headers, 1):
            max_length = len(header)
            col = get_column_letter(col_idx)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                if row[0].value:
                    max_length = max(max_length, len(str(row[0].value)))
            adjusted_width = min(max_length + 2, 50)  # Ограничение на 50 символов
            ws.column_dimensions[col].width = adjusted_width

        # Сохраняем файл
        try:
            wb.save(excel_path)
            self.logger.success(f"✅ Экспорт завершён: {excel_path}")
            return excel_path
        except Exception as e:
            self.logger.error(f"Ошибка сохранения Excel-файла: {e}")
            raise